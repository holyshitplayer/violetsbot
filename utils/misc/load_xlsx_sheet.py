from openpyxl import load_workbook

from data.config import KIDS_FILE, STARTERS_FILE, ORDERS_FILE


def load_kids_sheet():
    return load_workbook(KIDS_FILE).active


def load_starters_sheet():
    return load_workbook(STARTERS_FILE).active


def load_orders_sheet():
    return load_workbook(ORDERS_FILE).active
