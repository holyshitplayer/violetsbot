import openpyxl
from openpyxl.utils import get_column_letter

from utils.db_api import Database
from utils.misc.load_xlsx_sheet import load_starters_sheet, load_kids_sheet


def sorts_values_from_excel():
    sorts = set()
    for row in load_kids_sheet()[f"A2:A{load_kids_sheet().max_row}"]:
        for cell in row:
            sorts.add(cell.value)
    for row in load_starters_sheet()[f"A2:A{load_starters_sheet().max_row}"]:
        for cell in row:
            sorts.add(cell.value)
    values = list(zip(iter(sorted(list(sorts)))))
    return values


def content_values_from_excel(db: Database, sheet: openpyxl.worksheet.worksheet.Worksheet):
    values = []

    max_column = sheet.max_column
    max_row = sheet.max_row
    for current_column in range(2, max_column + 1):
        column_letter = get_column_letter(current_column)
        for current_row in range(2, max_row + 1):
            cell_value = sheet[f"{column_letter}{current_row}"].value
            if cell_value > 0 or cell_value is None:
                list_title = sheet[f"{column_letter}1"].value
                if "#" in list_title:
                    list_title = list_title.replace("#", "№")
                sort_name = sheet[f"A{current_row}"].value
                sort_id = db.select_sort(name=f"{sort_name}")[0]
                sort_count = cell_value
                values.append((sort_id, list_title, sort_count))

    return values
